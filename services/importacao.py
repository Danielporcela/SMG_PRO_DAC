"""Importação de cadastros a partir de planilhas Excel.

A ideia é carregar a frota, os motoristas, as oficinas e o estoque inicial sem
digitar item por item. O fluxo é sempre o mesmo:

1. o usuário baixa o modelo da planilha (já com os cabeçalhos certos);
2. envia a planilha preenchida e o sistema mostra uma prévia com os erros;
3. só depois de conferir é que os dados são gravados.

Nada é gravado pela metade: se a gravação falhar, a transação é desfeita.
"""
import io
from datetime import date

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from extensions import db
from models import Fornecedor, Motorista, Peca, Veiculo
from services.calculos import movimentar_estoque
from services.crud import ErroNegocio
from services.tempo import ler_data

# Cada modelo declara suas colunas: (cabeçalho, campo, tipo, obrigatório, exemplo)
MODELOS = {
    "veiculos": {
        "titulo": "Veículos",
        "modelo": Veiculo,
        "chave": "placa",
        "colunas": [
            ("Prefixo", "prefixo", "texto", True, "FR-101"),
            ("Placa", "placa", "placa", True, "ABC1D23"),
            ("Marca", "marca", "texto", False, "Mercedes-Benz"),
            ("Modelo", "modelo", "texto", False, "OF-1721"),
            ("Ano", "ano", "inteiro", False, 2019),
            ("Tipo", "tipo", "texto", False, "Ônibus"),
            ("Combustível", "combustivel", "texto", False, "Diesel S10"),
            ("Centro de custo", "centro_custo", "texto", False, "Transporte escolar"),
            ("Setor", "setor", "texto", False, "Operação"),
            ("Hodômetro", "hodometro", "numero", False, 120000),
            ("Km última troca de óleo", "km_ultima_troca_oleo", "numero", False, 118000),
            ("Intervalo de troca (km)", "intervalo_troca_oleo", "numero", False, 10000),
            ("Última preventiva", "data_ultima_preventiva", "data", False, "2026-05-20"),
            ("Intervalo preventiva (dias)", "intervalo_preventiva_dias", "inteiro", False, 90),
            ("Orçamento mensal", "orcamento_mensal", "numero", False, 5000),
        ],
    },
    "motoristas": {
        "titulo": "Motoristas",
        "modelo": Motorista,
        "chave": None,
        "colunas": [
            ("Nome", "nome", "texto", True, "João da Silva"),
            ("Matrícula", "matricula", "texto", False, "1234"),
            ("CNH", "cnh", "texto", False, "01234567890"),
            ("Categoria", "categoria_cnh", "texto", False, "D"),
            ("Validade da CNH", "validade_cnh", "data", False, "2028-03-15"),
            ("Telefone", "telefone", "texto", False, "(00) 90000-0000"),
            ("Setor", "setor", "texto", False, "Operação"),
        ],
    },
    "fornecedores": {
        "titulo": "Oficinas, postos e fornecedores",
        "modelo": Fornecedor,
        "chave": None,
        "colunas": [
            ("Nome", "nome", "texto", True, "Oficina Central"),
            ("Tipo", "tipo", "texto", False, "Oficina"),
            ("CNPJ", "cnpj", "texto", False, "00.000.000/0001-00"),
            ("Telefone", "telefone", "texto", False, "(00) 0000-0000"),
            ("Contato", "contato", "texto", False, "Carlos"),
            ("Cidade", "cidade", "texto", False, "Sua Cidade"),
        ],
    },
    "pecas": {
        "titulo": "Estoque de peças",
        "modelo": Peca,
        "chave": "codigo",
        "colunas": [
            ("Código", "codigo", "texto", True, "FIL-001"),
            ("Descrição", "descricao", "texto", True, "Filtro de óleo motor"),
            ("Grupo", "grupo", "texto", False, "Motor"),
            ("Unidade", "unidade", "texto", False, "UN"),
            ("Saldo inicial", "quantidade_inicial", "numero", False, 20),
            ("Custo unitário", "custo_unitario", "numero", False, 48.90),
            ("Estoque mínimo", "estoque_minimo", "numero", False, 5),
            ("Localização", "localizacao", "texto", False, "Prateleira A1"),
        ],
    },
}

LIMITE_LINHAS = 5000


def gerar_modelo(tipo):
    """Monta o arquivo .xlsx que o usuário baixa para preencher."""
    config = _config(tipo)
    wb = Workbook()
    ws = wb.active
    ws.title = config["titulo"][:31]

    cabecalhos = [c[0] for c in config["colunas"]]
    ws.append(cabecalhos)
    ws.append([c[4] for c in config["colunas"]])   # linha de exemplo

    fundo = PatternFill("solid", fgColor="0F3D56")
    for indice, coluna in enumerate(config["colunas"], start=1):
        celula = ws.cell(row=1, column=indice)
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = fundo
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        if coluna[3]:
            celula.value = f"{coluna[0]} *"
        ws.column_dimensions[get_column_letter(indice)].width = max(len(coluna[0]) + 6, 14)
    for celula in ws[2]:
        celula.font = Font(italic=True, color="8A8A8A")
    ws.freeze_panes = "A2"

    ws.append([])
    ws.append(["Colunas com * são obrigatórias. Apague a linha de exemplo antes de enviar."])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=9, color="666666")

    saida = io.BytesIO()
    wb.save(saida)
    saida.seek(0)
    return saida


def _config(tipo):
    if tipo not in MODELOS:
        raise ErroNegocio("Tipo de importação desconhecido.")
    return MODELOS[tipo]


def _converter(valor, tipo, cabecalho):
    if valor is None or str(valor).strip() == "":
        return None
    texto = str(valor).strip()
    try:
        if tipo == "inteiro":
            return int(float(texto.replace(".", "").replace(",", ".")))
        if tipo == "numero":
            return float(texto.replace(" ", "").replace(".", "").replace(",", ".")) \
                if texto.count(",") == 1 else float(texto.replace(" ", ""))
        if tipo == "data":
            return ler_data(texto, cabecalho)
        if tipo == "placa":
            return texto.upper().replace("-", "").replace(" ", "")
    except (ValueError, TypeError, ErroNegocio):
        raise ValueError(f"'{cabecalho}' com valor inválido: {texto}")
    return texto


def ler_planilha(tipo, arquivo):
    """Lê a planilha e devolve as linhas já conferidas, com os erros apontados."""
    config = _config(tipo)
    try:
        wb = load_workbook(arquivo, data_only=True, read_only=True)
    except Exception:
        raise ErroNegocio("Não consegui abrir a planilha. Envie um arquivo .xlsx.")

    ws = wb.active
    linhas_brutas = list(ws.iter_rows(values_only=True))
    if not linhas_brutas:
        raise ErroNegocio("A planilha está vazia.")

    cabecalho = [str(c).replace("*", "").strip().lower() if c else ""
                 for c in linhas_brutas[0]]
    indices = {}
    faltando = []
    for titulo, campo, tipo_campo, obrigatorio, _ in config["colunas"]:
        chave = titulo.lower()
        if chave in cabecalho:
            indices[campo] = cabecalho.index(chave)
        elif obrigatorio:
            faltando.append(titulo)
    if faltando:
        raise ErroNegocio("A planilha não tem as colunas: " + ", ".join(faltando) +
                          ". Baixe o modelo e use os mesmos cabeçalhos.")

    existentes = set()
    if config["chave"]:
        coluna = getattr(config["modelo"], config["chave"])
        existentes = {str(v[0]).upper() for v in db.session.query(coluna).all() if v[0]}

    vistos = set()
    prontas, problemas = [], []

    for numero, bruta in enumerate(linhas_brutas[1:], start=2):
        if numero - 1 > LIMITE_LINHAS:
            problemas.append({"linha": numero, "erro":
                              f"A planilha passa de {LIMITE_LINHAS} linhas. Divida em partes."})
            break
        if not any(c is not None and str(c).strip() != "" for c in bruta):
            continue

        registro, erros = {}, []
        for titulo, campo, tipo_campo, obrigatorio, _ in config["colunas"]:
            if campo not in indices:
                continue
            posicao = indices[campo]
            valor = bruta[posicao] if posicao < len(bruta) else None
            try:
                convertido = _converter(valor, tipo_campo, titulo)
            except ValueError as e:
                erros.append(str(e))
                continue
            if obrigatorio and convertido in (None, ""):
                erros.append(f"'{titulo}' é obrigatório")
            registro[campo] = convertido

        if config["chave"]:
            chave = str(registro.get(config["chave"]) or "").upper()
            if chave and chave in existentes:
                erros.append(f"já existe no sistema ({chave})")
            elif chave and chave in vistos:
                erros.append(f"repetido na própria planilha ({chave})")
            elif chave:
                vistos.add(chave)

        if tipo == "veiculos" and not erros:
            from services.grupos_consumo import nome_grupo_consumo_legado
            candidato = Veiculo(**registro)
            grupo_consumo = nome_grupo_consumo_legado(candidato)
            if grupo_consumo:
                erros.append(
                    f"{grupo_consumo} é grupo de consumo. Use o menu Grupos de consumo.")

        # As datas saem como texto ISO para atravessar o JSON da prévia e
        # voltarem íntegras na hora de gravar.
        registro = {k: (v.isoformat() if isinstance(v, date) else v)
                    for k, v in registro.items()}

        if erros:
            problemas.append({"linha": numero, "erro": "; ".join(erros),
                              "dados": {k: str(v) for k, v in registro.items() if v is not None}})
        else:
            prontas.append({"linha": numero, "dados": registro})

    return {"tipo": tipo, "titulo": config["titulo"],
            "colunas": [c[0] for c in config["colunas"] if c[1] in indices],
            "campos": [c[1] for c in config["colunas"] if c[1] in indices],
            "prontas": prontas, "problemas": problemas,
            "total": len(prontas) + len(problemas)}


def gravar(tipo, linhas):
    """Grava as linhas aprovadas. Ou entra tudo, ou não entra nada."""
    config = _config(tipo)
    Model = config["modelo"]
    if not linhas:
        raise ErroNegocio("Não há linhas válidas para importar.")

    tipos = {campo: (titulo, tipo_campo)
             for titulo, campo, tipo_campo, _, _ in config["colunas"]}
    gravadas = 0
    try:
        for item in linhas:
            bruto = dict(item.get("dados") or item)
            # Reconverte o que veio da tela: o JSON não guarda datas nem números.
            dados = {}
            for campo, valor in bruto.items():
                if campo not in tipos:
                    continue
                titulo, tipo_campo = tipos[campo]
                dados[campo] = _converter(valor, tipo_campo, titulo)
            saldo = dados.pop("quantidade_inicial", None)
            custo = dados.get("custo_unitario") or 0
            obj = Model(**dados)
            if tipo == "veiculos":
                from services.grupos_consumo import nome_grupo_consumo_legado
                grupo_consumo = nome_grupo_consumo_legado(obj)
                if grupo_consumo:
                    raise ErroNegocio(
                        f"{grupo_consumo} é grupo de consumo. Use o menu Grupos de consumo.")
            db.session.add(obj)
            db.session.flush()
            if tipo == "pecas" and saldo:
                movimentar_estoque(obj.id, "entrada", saldo, custo,
                                   documento="Importação de planilha")
            gravadas += 1
        db.session.commit()
    except ErroNegocio:
        db.session.rollback()
        raise
    except Exception as e:
        db.session.rollback()
        raise ErroNegocio(f"A importação foi cancelada e nada foi gravado "
                          f"({e.__class__.__name__}). Confira a planilha e tente de novo.")
    return gravadas
