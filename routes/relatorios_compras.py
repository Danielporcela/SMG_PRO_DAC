"""Relatório de ordens de compra — planilha para o financeiro.

Uma rota só, no mesmo endereço das outras exportações do sistema:

    /relatorios/ordens_compra.xlsx?inicio=2026-01-01&fim=2026-01-31&status=Pendente

A planilha sai com três abas:

    Resumo   — quanto tem em cada situação, por setor e por prioridade
    Ordens   — uma linha por ordem, com a trilha de aprovação
    Itens    — uma linha por item, para conferir o que foi pedido e cotar

Quem enxerga: quem tem 'visualizar' em Ordens de compra OU em Relatórios —
o financeiro costuma ter só a segunda, e o almoxarifado só a primeira.
"""
from datetime import datetime
from io import BytesIO

from flask import Blueprint, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from models import OrdemCompra
from services.crud import checar_tela, nivel_permite

# ATENÇÃO: esta blueprint NÃO leva url_prefix="/api" — as rotas já começam
# em /relatorios, igual às outras exportações do sistema.
bp_relatorios_compras = Blueprint("relatorios_compras", __name__)

MOEDA = 'R$ #,##0.00'
PETROLEO = "0F3D56"
CINZA = "F5F7F9"

CORES_SITUACAO = {"Pendente": "FFF3CD", "Aprovada": "D9F2DE",
                  "Reprovada": "FADBD8", "Comprada": "DCE9F5"}


# ---------------------------------------------------------------- auxiliares
def _permissao():
    """Devolve a resposta 401/403 pronta, ou None quando pode seguir."""
    if nivel_permite("compras", "visualizar") or nivel_permite("relatorios", "visualizar"):
        return None
    return checar_tela("compras", "visualizar")


def _data(texto):
    try:
        return datetime.strptime(str(texto)[:10], "%Y-%m-%d").date()
    except (ValueError, TypeError):
        return None


def _consultar():
    inicio, fim = _data(request.args.get("inicio")), _data(request.args.get("fim"))
    status = (request.args.get("status") or "").strip()
    q = OrdemCompra.query
    if inicio:
        q = q.filter(OrdemCompra.data_solicitacao >= inicio)
    if fim:
        q = q.filter(OrdemCompra.data_solicitacao <= fim)
    if status:
        q = q.filter(OrdemCompra.status == status)
    return q.order_by(OrdemCompra.data_solicitacao.desc(), OrdemCompra.id.desc()).all(), inicio, fim, status


def _cabecalho(aba, titulos, linha=1):
    fundo = PatternFill("solid", fgColor=PETROLEO)
    borda = Border(bottom=Side(style="thin", color="D3DBE2"))
    for coluna, titulo in enumerate(titulos, start=1):
        celula = aba.cell(row=linha, column=coluna, value=titulo)
        celula.font = Font(bold=True, color="FFFFFF", size=10)
        celula.fill = fundo
        celula.border = borda
        celula.alignment = Alignment(vertical="center", wrap_text=True)
    aba.row_dimensions[linha].height = 26


def _ajustar_larguras(aba, larguras):
    for coluna, largura in enumerate(larguras, start=1):
        aba.column_dimensions[get_column_letter(coluna)].width = largura


def _titulo_bloco(aba, linha, texto):
    celula = aba.cell(row=linha, column=1, value=texto)
    celula.font = Font(bold=True, size=11, color=PETROLEO)
    return linha + 1


# -------------------------------------------------------------- aba Resumo
def _montar_resumo(aba, ordens, inicio, fim, status):
    aba.sheet_view.showGridLines = False
    _ajustar_larguras(aba, [30, 14, 18, 18, 16])

    aba["A1"] = "Relatório de ordens de compra"
    aba["A1"].font = Font(bold=True, size=14, color=PETROLEO)
    periodo = "todo o histórico"
    if inicio or fim:
        periodo = f"{inicio.strftime('%d/%m/%Y') if inicio else 'início'} a " \
                  f"{fim.strftime('%d/%m/%Y') if fim else 'hoje'}"
    aba["A2"] = f"Período: {periodo}" + (f" · Situação: {status}" if status else "")
    aba["A2"].font = Font(size=10, color="6B7C8A")
    aba["A3"] = f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')} · " \
                "valores estimados pela operação — a ordem de compra não movimenta estoque"
    aba["A3"].font = Font(size=9, italic=True, color="6B7C8A")

    linha = _titulo_bloco(aba, 5, "Por situação")
    _cabecalho(aba, ["Situação", "Ordens", "Valor estimado", "% do valor"], linha)
    linha += 1

    total_valor = sum(o.valor_total for o in ordens) or 0
    for situacao in ("Pendente", "Aprovada", "Reprovada", "Comprada"):
        do_grupo = [o for o in ordens if o.status == situacao]
        valor = sum(o.valor_total for o in do_grupo)
        aba.cell(row=linha, column=1, value=situacao).fill = \
            PatternFill("solid", fgColor=CORES_SITUACAO[situacao])
        aba.cell(row=linha, column=2, value=len(do_grupo))
        aba.cell(row=linha, column=3, value=valor).number_format = MOEDA
        celula = aba.cell(row=linha, column=4, value=(valor / total_valor) if total_valor else 0)
        celula.number_format = '0.0%'
        linha += 1

    aba.cell(row=linha, column=1, value="Total").font = Font(bold=True)
    aba.cell(row=linha, column=2, value=len(ordens)).font = Font(bold=True)
    celula = aba.cell(row=linha, column=3, value=total_valor)
    celula.number_format = MOEDA
    celula.font = Font(bold=True)
    linha += 3

    def bloco_por(rotulo, chave):
        nonlocal linha
        grupos = {}
        for o in ordens:
            nome = (getattr(o, chave) or "— não informado —")
            atual = grupos.setdefault(nome, [0, 0.0])
            atual[0] += 1
            atual[1] += o.valor_total
        linha = _titulo_bloco(aba, linha, rotulo)
        _cabecalho(aba, [rotulo.replace("Por ", "").capitalize(), "Ordens", "Valor estimado"], linha)
        linha += 1
        for nome, (quantidade, valor) in sorted(grupos.items(), key=lambda x: -x[1][1]):
            aba.cell(row=linha, column=1, value=nome)
            aba.cell(row=linha, column=2, value=quantidade)
            aba.cell(row=linha, column=3, value=valor).number_format = MOEDA
            linha += 1
        linha += 2

    bloco_por("Por setor", "setor")
    bloco_por("Por prioridade", "prioridade")

    # Itens mais pedidos — ajuda a enxergar o que vale negociar por volume
    contagem = {}
    for o in ordens:
        for i in o.itens:
            chave = (i.peca.descricao if i.peca else i.descricao) or "—"
            atual = contagem.setdefault(chave, [0.0, 0.0, set()])
            atual[0] += i.quantidade or 0
            atual[1] += i.subtotal
            atual[2].add(o.id)
    if contagem:
        linha = _titulo_bloco(aba, linha, "Itens mais pedidos")
        _cabecalho(aba, ["Item", "Qtde somada", "Valor estimado", "Em quantas ordens"], linha)
        linha += 1
        for nome, (quantidade, valor, ordens_do_item) in \
                sorted(contagem.items(), key=lambda x: -x[1][1])[:20]:
            aba.cell(row=linha, column=1, value=nome)
            aba.cell(row=linha, column=2, value=quantidade)
            aba.cell(row=linha, column=3, value=valor).number_format = MOEDA
            aba.cell(row=linha, column=4, value=len(ordens_do_item))
            linha += 1


# ------------------------------------------------------------- aba Ordens
def _montar_ordens(aba, ordens):
    titulos = ["OC", "Solicitada em", "Solicitante", "Setor", "Prioridade", "Situação",
               "Fornecedor sugerido", "Itens", "Valor estimado", "Aprovada/reprovada por",
               "Data da decisão", "Motivo da reprovação", "Comprada por", "Data da compra",
               "Justificativa"]
    _cabecalho(aba, titulos)
    _ajustar_larguras(aba, [11, 14, 20, 16, 12, 12, 24, 8, 16, 22, 15, 34, 20, 14, 44])

    for posicao, o in enumerate(ordens, start=2):
        valores = [o.numero, o.data_solicitacao, o.solicitante, o.setor, o.prioridade, o.status,
                   o.fornecedor.nome if o.fornecedor else None, len(o.itens), o.valor_total,
                   o.aprovado_por, o.data_aprovacao, o.motivo_reprovacao,
                   o.comprado_por, o.data_compra, o.justificativa]
        for coluna, valor in enumerate(valores, start=1):
            aba.cell(row=posicao, column=coluna, value=valor)
        aba.cell(row=posicao, column=2).number_format = 'DD/MM/YYYY'
        aba.cell(row=posicao, column=9).number_format = MOEDA
        aba.cell(row=posicao, column=11).number_format = 'DD/MM/YYYY'
        aba.cell(row=posicao, column=14).number_format = 'DD/MM/YYYY'
        aba.cell(row=posicao, column=6).fill = \
            PatternFill("solid", fgColor=CORES_SITUACAO.get(o.status, "FFFFFF"))
        aba.cell(row=posicao, column=15).alignment = Alignment(wrap_text=True, vertical="top")

    aba.freeze_panes = "A2"
    if ordens:
        aba.auto_filter.ref = f"A1:{get_column_letter(len(titulos))}{len(ordens) + 1}"


# -------------------------------------------------------------- aba Itens
def _montar_itens(aba, ordens):
    titulos = ["OC", "Situação", "Solicitada em", "Setor", "Item", "Origem", "Código da peça",
               "Saldo atual", "Unidade", "Qtde", "Valor unit. estimado", "Total", "Observação"]
    _cabecalho(aba, titulos)
    _ajustar_larguras(aba, [11, 12, 14, 16, 40, 11, 15, 12, 9, 10, 18, 14, 32])

    linha = 2
    for o in ordens:
        for i in o.itens:
            valores = [o.numero, o.status, o.data_solicitacao, o.setor, i.descricao,
                       "Estoque" if i.peca_id else "Digitado",
                       i.peca.codigo if i.peca else None,
                       i.peca.quantidade if i.peca else None,
                       i.unidade, i.quantidade, i.valor_unitario, i.subtotal, i.observacao]
            for coluna, valor in enumerate(valores, start=1):
                aba.cell(row=linha, column=coluna, value=valor)
            aba.cell(row=linha, column=3).number_format = 'DD/MM/YYYY'
            aba.cell(row=linha, column=11).number_format = MOEDA
            aba.cell(row=linha, column=12).number_format = MOEDA
            aba.cell(row=linha, column=2).fill = \
                PatternFill("solid", fgColor=CORES_SITUACAO.get(o.status, "FFFFFF"))
            linha += 1

    aba.freeze_panes = "A2"
    if linha > 2:
        aba.auto_filter.ref = f"A1:{get_column_letter(len(titulos))}{linha - 1}"
        total = aba.cell(row=linha, column=11, value="Total")
        total.font = Font(bold=True)
        total.alignment = Alignment(horizontal="right")
        soma = aba.cell(row=linha, column=12, value=f"=SUM(L2:L{linha - 1})")
        soma.font = Font(bold=True)
        soma.number_format = MOEDA
        soma.fill = PatternFill("solid", fgColor=CINZA)


# ------------------------------------------------------------------- rota
@bp_relatorios_compras.get("/relatorios/ordens_compra.xlsx")
def ordens_compra_xlsx():
    negado = _permissao()
    if negado:
        return negado

    ordens, inicio, fim, status = _consultar()

    planilha = Workbook()
    _montar_resumo(planilha.active, ordens, inicio, fim, status)
    planilha.active.title = "Resumo"
    _montar_ordens(planilha.create_sheet("Ordens"), ordens)
    _montar_itens(planilha.create_sheet("Itens"), ordens)

    arquivo = BytesIO()
    planilha.save(arquivo)
    arquivo.seek(0)
    nome = f"ordens_compra_{datetime.now().strftime('%Y-%m-%d')}.xlsx"
    return send_file(arquivo, as_attachment=True, download_name=nome,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
