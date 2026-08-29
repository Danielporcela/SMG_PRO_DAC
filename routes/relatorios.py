"""Exportação de relatórios em PDF, Excel e CSV com filtros de período."""
import csv
import io
from datetime import date, datetime, timedelta

from flask import Blueprint, jsonify, request, send_file
from openpyxl import Workbook
from openpyxl.styles import (Alignment, Border, Font, PatternFill, Side)
from openpyxl.styles import Alignment as openpyxl_align
from openpyxl.styles import Border as openpyxl_border
from openpyxl.styles import Font as openpyxl_font
from openpyxl.styles import Side as openpyxl_side
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from extensions import db
from models import (Abastecimento, ItemOS, MovimentoEstoque, NotaFiscal, OrdemServico,
                    Peca, Pneu, Veiculo)
from services import indicadores
from services.crud import login_obrigatorio, perfil_obrigatorio, registrar_log, visualizar_tela
from services.restauracao import restaurar
from services.tempo import agora, hoje, ler_data

bp_relatorios = Blueprint("relatorios", __name__, url_prefix="/relatorios")

TITULOS = {
    "abastecimentos": "Abastecimentos",
    "manutencoes": "Ordens de serviço",
    "veiculos": "Frota cadastrada",
    "pneus": "Controle de pneus",
    "pneus_movimentos": "Movimentação de pneus",
    "estoque": "Posição de estoque",
    "movimentos": "Movimentação de estoque",
    "lubrificantes": "Óleos e fluidos",
    "custos": "Custos por veículo",
    "gastos_nf": "Gastos com notas fiscais",
}


def _periodo():
    inicio = ler_data(request.args.get("inicio"), "início do período")
    fim = ler_data(request.args.get("fim"), "fim do período")
    return inicio or hoje() - timedelta(days=29), fim or hoje()


def montar_dados(relatorio):
    """Retorna (cabecalhos, linhas) já formatados para qualquer formato de saída."""
    inicio, fim = _periodo()
    veiculo_id = request.args.get("veiculo_id", type=int)
    motorista_id = request.args.get("motorista_id", type=int)
    fornecedor_id = request.args.get("fornecedor_id", type=int)
    centro_custo = request.args.get("centro_custo")
    grupo = request.args.get("grupo")
    status_peca = (request.args.get("status_peca") or "inventario").strip().lower()

    if relatorio == "abastecimentos":
        q = Abastecimento.query.filter(Abastecimento.data.between(inicio, fim))
        if veiculo_id:
            q = q.filter(Abastecimento.veiculo_id == veiculo_id)
        if motorista_id:
            q = q.filter(Abastecimento.motorista_id == motorista_id)
        if fornecedor_id:
            q = q.filter(Abastecimento.fornecedor_id == fornecedor_id)
        cab = ["Data", "Veículo", "Motorista", "Posto", "Km atual", "Km rodados",
               "Litros", "R$/L", "Km/L", "Custo/km", "Total R$"]
        linhas = [[a.data.strftime("%d/%m/%Y"),
                   f"{a.veiculo.prefixo}/{a.veiculo.placa}" if a.veiculo else "—",
                   a.motorista.nome if a.motorista else "—",
                   a.fornecedor.nome if a.fornecedor else "—",
                   round(a.km_atual or 0), round(a.km_percorridos or 0),
                   round(a.litros or 0, 2), round(a.valor_litro or 0, 3),
                   round(a.km_por_litro or 0, 2), round(a.custo_por_km or 0, 3),
                   round(a.valor_total or 0, 2)]
                  for a in q.order_by(Abastecimento.data).all()]

    elif relatorio == "manutencoes":
        q = OrdemServico.query.filter(OrdemServico.data_abertura.between(inicio, fim))
        if veiculo_id:
            q = q.filter(OrdemServico.veiculo_id == veiculo_id)
        if fornecedor_id:
            q = q.filter(OrdemServico.fornecedor_id == fornecedor_id)
        cab = ["OS", "Abertura", "Fechamento", "Veículo", "Tipo", "Grupo", "Status",
               "Oficina", "Dias parado", "Peças R$", "Mão de obra R$", "Serviços R$", "Total R$"]
        linhas = [[o.numero, o.data_abertura.strftime("%d/%m/%Y") if o.data_abertura else "",
                   o.data_fechamento.strftime("%d/%m/%Y") if o.data_fechamento else "—",
                   f"{o.veiculo.prefixo}/{o.veiculo.placa}" if o.veiculo else "—",
                   o.tipo, o.grupo or "—", o.status,
                   o.fornecedor.nome if o.fornecedor else "—", o.dias_parado,
                   o.custo_pecas, round(o.custo_mao_obra or 0, 2),
                   o.custo_servicos_total, o.custo_total]
                  for o in q.order_by(OrdemServico.data_abertura).all()]

    elif relatorio == "veiculos":
        q = Veiculo.query
        if centro_custo:
            q = q.filter(Veiculo.centro_custo == centro_custo)
        if veiculo_id:
            q = q.filter(Veiculo.id == veiculo_id)
        cab = ["Prefixo", "Placa", "Marca/Modelo", "Ano", "Tipo", "Centro de custo",
               "Setor", "Hodômetro", "Situação", "Próx. troca óleo (km)"]
        linhas = [[v.prefixo, v.placa, f"{v.marca or ''} {v.modelo or ''}".strip(), v.ano or "",
                   v.tipo or "", v.centro_custo or "", v.setor or "", round(v.hodometro or 0),
                   v.situacao, round(v.km_proxima_troca_oleo)]
                  for v in q.order_by(Veiculo.prefixo).all()]

    elif relatorio == "pneus":
        q = Pneu.query
        if veiculo_id:
            q = q.filter(Pneu.veiculo_id == veiculo_id)
        cab = ["Nº fogo", "Veículo", "Posição", "Marca", "Medida", "Sulco (mm)",
               "Vida", "Km rodados", "Status", "Situação"]
        linhas = []
        for p in q.order_by(Pneu.numero_fogo).all():
            d = p.to_dict()
            linhas.append([p.numero_fogo, d["veiculo_nome"] or "—", p.posicao or "—",
                           p.marca or "—", p.medida or "—", round(p.sulco_mm or 0, 1),
                           p.vida, d["km_rodados"], p.status,
                           "TROCAR" if d["trocar"] else "OK"])

    elif relatorio == "pneus_movimentos":
        # Cada troca registrada nas ordens de serviço: quando a posição do pneu
        # é informada na OS, o sistema guarda a posição e qual pneu foi baixado
        # (pneu_substituido_id). Este relatório lista essas trocas por período.
        q = (ItemOS.query
             .join(OrdemServico, ItemOS.ordem_servico_id == OrdemServico.id)
             .filter(ItemOS.posicao_pneu.isnot(None),
                     OrdemServico.data_abertura.between(inicio, fim)))
        if veiculo_id:
            q = q.filter(OrdemServico.veiculo_id == veiculo_id)
        cab = ["Data", "OS", "Veículo", "Posição", "Pneu retirado",
               "Sulco retirado (mm)", "Pneu instalado", "Item aplicado na OS"]
        linhas = []
        for item in q.order_by(OrdemServico.data_abertura, ItemOS.id).all():
            os_obj = item.ordem
            antigo = (db.session.get(Pneu, item.pneu_substituido_id)
                      if item.pneu_substituido_id else None)
            novo = None
            if os_obj and os_obj.veiculo_id:
                q_novo = Pneu.query.filter(Pneu.veiculo_id == os_obj.veiculo_id,
                                           Pneu.posicao == item.posicao_pneu)
                if antigo:
                    q_novo = q_novo.filter(Pneu.id != antigo.id)
                if os_obj.data_abertura:
                    q_novo = q_novo.filter(Pneu.data_instalacao >= os_obj.data_abertura)
                novo = q_novo.order_by(Pneu.data_instalacao, Pneu.id).first()
            linhas.append([
                os_obj.data_abertura.strftime("%d/%m/%Y")
                if os_obj and os_obj.data_abertura else "—",
                os_obj.numero if os_obj else "—",
                f"{os_obj.veiculo.prefixo}/{os_obj.veiculo.placa}"
                if os_obj and os_obj.veiculo else "—",
                item.posicao_pneu,
                antigo.numero_fogo if antigo else "—",
                round(antigo.sulco_mm or 0, 1) if antigo else "—",
                novo.numero_fogo if novo else "—",
                item.descricao or (item.peca.descricao if item.peca else "Pneu")])

    elif relatorio == "estoque":
        q = Peca.query
        if grupo:
            q = q.filter(Peca.grupo == grupo)

        saldo = db.func.coalesce(Peca.quantidade, 0)
        minimo = db.func.coalesce(Peca.estoque_minimo, 0)

        if status_peca == "repor":
            q = q.filter(saldo <= minimo)
        elif status_peca == "ok":
            q = q.filter(saldo > minimo)

        cab = ["Código", "Descrição", "Grupo", "Un.", "Saldo", "Mínimo",
               "Custo unit. R$", "Valor total R$", "Situação"]
        linhas = [[p.codigo, p.descricao, p.grupo or "—", p.unidade,
                   round(p.quantidade or 0, 2), round(p.estoque_minimo or 0, 2),
                   round(p.custo_unitario or 0, 2),
                   round((p.quantidade or 0) * (p.custo_unitario or 0), 2),
                   "REPOR" if (p.quantidade or 0) <= (p.estoque_minimo or 0) else "OK"]
                  for p in q.order_by(Peca.grupo, Peca.descricao).all()]

    elif relatorio == "movimentos":
        q = MovimentoEstoque.query.filter(MovimentoEstoque.data.between(inicio, fim))
        if grupo:
            q = q.join(Peca, MovimentoEstoque.peca_id == Peca.id).filter(Peca.grupo == grupo)
        cab = ["Data", "Peça", "Tipo", "Quantidade", "Custo unit. R$", "Total R$", "Documento"]
        linhas = [[m.data.strftime("%d/%m/%Y"),
                   f"{m.peca.codigo} - {m.peca.descricao}" if m.peca else "—",
                   m.tipo.capitalize(), round(m.quantidade or 0, 2),
                   round(m.custo_unitario or 0, 2),
                   round((m.quantidade or 0) * (m.custo_unitario or 0), 2),
                   m.documento or "—"]
                  for m in q.order_by(MovimentoEstoque.data).all()]

    elif relatorio == "lubrificantes":
        # Entradas e saídas de óleos, fluidos (hidráulico, freio, ATF), graxas e
        # lubrificantes em geral. O produto é reconhecido pelo texto do cadastro
        # da peça (descrição, grupo ou código), no mesmo espírito do
        # reconhecimento de pneus feito em services/correcoes_os.py.
        TERMOS_LUBRIFICANTES = ("oleo", "óleo", "fluido", "fluído", "hidraulic",
                                "hidráulic", "lubrific", "graxa", "atf")

        def _eh_lubrificante(peca):
            if peca is None:
                return False
            texto = " ".join(str(v or "") for v in
                             (peca.descricao, peca.grupo, peca.codigo)).casefold()
            return any(t in texto for t in TERMOS_LUBRIFICANTES)

        q = MovimentoEstoque.query.filter(MovimentoEstoque.data.between(inicio, fim))
        cab = ["Data", "Produto", "Grupo", "Tipo", "Quantidade", "Un.",
               "Custo unit. R$", "Total R$", "Veículo / OS", "Documento"]
        linhas = []
        for m in q.order_by(MovimentoEstoque.data).all():
            if not _eh_lubrificante(m.peca):
                continue
            os_obj = (db.session.get(OrdemServico, m.ordem_servico_id)
                      if m.ordem_servico_id else None)
            if veiculo_id and (os_obj is None or os_obj.veiculo_id != veiculo_id):
                continue
            destino = "—"
            if os_obj:
                veic = (f"{os_obj.veiculo.prefixo}/{os_obj.veiculo.placa}"
                        if os_obj.veiculo else "")
                destino = f"{veic} · {os_obj.numero}".strip(" ·")
            linhas.append([m.data.strftime("%d/%m/%Y"),
                           f"{m.peca.codigo} - {m.peca.descricao}" if m.peca else "—",
                           (m.peca.grupo or "—") if m.peca else "—",
                           (m.tipo or "").capitalize(),
                           round(m.quantidade or 0, 2),
                           m.peca.unidade if m.peca else "",
                           round(m.custo_unitario or 0, 2),
                           round((m.quantidade or 0) * (m.custo_unitario or 0), 2),
                           destino, m.documento or "—"])

    elif relatorio == "custos":
        dados = indicadores.series_graficos(inicio.isoformat(), fim.isoformat())["por_veiculo"]
        cab = ["Veículo", "Placa", "Km rodados", "Combustível R$", "Manutenção R$",
               "Total R$", "Custo/km R$", "Km/L", "Orçamento R$"]
        linhas = [[d["veiculo"], d["placa"], d["km"], d["combustivel"], d["manutencao"],
                   d["total"], d["custo_km"], d["consumo"], d["orcamento"]] for d in dados]

    elif relatorio == "gastos_nf":
        # Gasto real de compra de peças (Módulo 11): só entra a nota já
        # finalizada (deu entrada de fato no estoque), contada pela data de
        # entrada — igual ao que o painel usa em "gasto_compras". "Valor
        # peças" é o mesmo valor que a tela de Estoque já mostra por nota;
        # "Tributos" soma ICMS/PIS/COFINS/IBS/CBS lançados nos itens, só
        # como referência fiscal.
        def _tributos(nota):
            return round(sum(nota._total_fiscal(campo) for campo in
                             ("valor_icms", "valor_pis", "valor_cofins",
                              "valor_ibs", "valor_cbs")), 2)

        q = NotaFiscal.query.filter(NotaFiscal.status == "Finalizada",
                                    NotaFiscal.data_entrada.between(inicio, fim))
        if fornecedor_id:
            q = q.filter(NotaFiscal.fornecedor_id == fornecedor_id)
        notas = q.order_by(NotaFiscal.data_entrada, NotaFiscal.id).all()
        cab = ["NF", "Emissão", "Entrada", "Fornecedor", "Itens",
               "Valor peças R$", "Tributos R$", "Total R$"]
        linhas = [[n.identificacao,
                   n.data_emissao.strftime("%d/%m/%Y") if n.data_emissao else "—",
                   n.data_entrada.strftime("%d/%m/%Y") if n.data_entrada else "—",
                   n.fornecedor.nome if n.fornecedor else "—", len(n.itens),
                   n.valor_total, _tributos(n), round(n.valor_total + _tributos(n), 2)]
                  for n in notas]
    else:
        cab, linhas = ["Relatório"], [["Relatório não encontrado."]]

    return cab, linhas, inicio, fim


def _rotulo_status_peca():
    status = (request.args.get("status_peca") or "inventario").strip().lower()
    return {
        "inventario": "Inventário completo",
        "repor": "Para repor",
        "ok": "Estoque OK",
        "grupo": "Por grupo",
    }.get(status, "Inventário completo")


def _descricao_filtros(relatorio):
    partes = []
    grupo = request.args.get("grupo")
    if relatorio == "estoque":
        partes.append(_rotulo_status_peca())
    if grupo:
        partes.append(f"Grupo: {grupo}")
    return " · ".join(partes)


@bp_relatorios.get("/<relatorio>.csv")
@visualizar_tela("relatorios")
def exportar_csv(relatorio):
    cab, linhas, *_ = montar_dados(relatorio)
    buffer = io.StringIO()
    escritor = csv.writer(buffer, delimiter=";")
    escritor.writerow(cab)
    escritor.writerows(linhas)
    dados = io.BytesIO(buffer.getvalue().encode("utf-8-sig"))
    return send_file(dados, mimetype="text/csv", as_attachment=True,
                     download_name=f"sgmf_{relatorio}_{hoje():%Y%m%d}.csv")


@bp_relatorios.get("/<relatorio>.xlsx")
@visualizar_tela("relatorios")
def exportar_excel(relatorio):
    cab, linhas, inicio, fim = montar_dados(relatorio)
    wb = Workbook()
    ws = wb.active
    ws.title = TITULOS.get(relatorio, "Relatório")[:31]

    descricao_filtros = _descricao_filtros(relatorio)
    ws.append([f"SGMF Pro · {TITULOS.get(relatorio, relatorio)}"
               + (f" · {descricao_filtros}" if descricao_filtros else "")])
    ws.append([f"Período: {inicio:%d/%m/%Y} a {fim:%d/%m/%Y} · "
               f"emitido em {agora():%d/%m/%Y %H:%M}"])
    ws.append([])
    ws.append(cab)
    cabecalho_linha = ws.max_row
    for linha in linhas:
        ws.append(linha)

    ws["A1"].font = Font(bold=True, size=14, color="0F3D56")
    ws["A2"].font = Font(size=9, color="666666")
    fundo = PatternFill("solid", fgColor="0F3D56")
    for celula in ws[cabecalho_linha]:
        celula.font = Font(bold=True, color="FFFFFF")
        celula.fill = fundo
        celula.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for coluna in ws.columns:
        largura = max((len(str(c.value)) for c in coluna if c.value), default=10)
        ws.column_dimensions[coluna[0].column_letter].width = min(max(largura + 2, 10), 42)
    ws.freeze_panes = ws.cell(row=cabecalho_linha + 1, column=1)

    saida = io.BytesIO()
    wb.save(saida)
    saida.seek(0)
    return send_file(saida, as_attachment=True,
                     download_name=f"sgmf_{relatorio}_{hoje():%Y%m%d}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@bp_relatorios.get("/<relatorio>.pdf")
@visualizar_tela("relatorios")
def exportar_pdf(relatorio):
    cab, linhas, inicio, fim = montar_dados(relatorio)
    saida = io.BytesIO()
    doc = SimpleDocTemplate(saida, pagesize=landscape(A4),
                            leftMargin=12 * mm, rightMargin=12 * mm,
                            topMargin=12 * mm, bottomMargin=12 * mm,
                            title=f"SGMF Pro - {TITULOS.get(relatorio, relatorio)}")
    estilos = getSampleStyleSheet()
    descricao_filtros = _descricao_filtros(relatorio)
    elementos = [
        Paragraph(f"<b>SGMF Pro</b> · {TITULOS.get(relatorio, relatorio)}"
                  + (f" · {descricao_filtros}" if descricao_filtros else ""), estilos["Title"]),
        Paragraph(f"Período de {inicio:%d/%m/%Y} a {fim:%d/%m/%Y} — emitido em "
                  f"{agora():%d/%m/%Y às %H:%M}", estilos["Normal"]),
        Spacer(1, 8),
    ]

    mensagem_vazia = ("Nenhuma peça encontrada para os filtros selecionados."
                     if relatorio == "estoque" else "Nenhum lançamento no período.")
    dados = [cab] + [[str(c) for c in linha] for linha in linhas] if linhas else \
            [cab, [mensagem_vazia] + [""] * (len(cab) - 1)]
    tabela = Table(dados, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0F3D56")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
        ("ALIGN", (1, 1), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#C9D2DA")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F2F5F7")]),
        ("TOPPADDING", (0, 0), (-1, -1), 3),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
    ]))
    elementos += [tabela, Spacer(1, 10),
                  Paragraph(f"{len(linhas)} registro(s) · Sistema de Gestão de Manutenção de Frotas",
                            estilos["Italic"])]
    doc.build(elementos)
    saida.seek(0)
    return send_file(saida, as_attachment=True, mimetype="application/pdf",
                     download_name=f"sgmf_{relatorio}_{hoje():%Y%m%d}.pdf")


@bp_relatorios.get("/mecanico/<mecanico>.pdf")
@visualizar_tela("relatorios")
def relatorio_mecanico_pdf(mecanico):
    """Relatório de produtividade de um mecânico em PDF."""
    dados = _montar_dados_mecanico(mecanico)
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4,
                            leftMargin=12*mm, rightMargin=12*mm,
                            topMargin=12*mm, bottomMargin=12*mm)
    estilos = getSampleStyleSheet()
    elementos = [
        Paragraph(f"Relatório de Serviços — {mecanico.title()}", estilos["Heading1"]),
        Paragraph(
            f"Período: {dados['inicio']:%d/%m/%Y} a {dados['fim']:%d/%m/%Y}  •  "
            f"{dados['total_os']} OS  •  {dados['horas_totais']}h trabalhadas  •  "
            f"Emitido em {agora():%d/%m/%Y %H:%M}",
            estilos["Normal"]
        ),
        Spacer(1, 5*mm),
    ]

    # Resumo por tipo
    resumo = [["Tipo", "Qtd OS", "Horas", "Custo total"]]
    for linha in dados["por_tipo"]:
        resumo.append([linha["tipo"], str(linha["qtd"]),
                       linha["horas"], f"R$ {linha['custo']:,.2f}".replace(",","X").replace(".",",").replace("X",".")])
    t = Table(resumo, colWidths=[60*mm, 30*mm, 30*mm, 50*mm])
    t.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#0F3D56")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 8.5),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F4F7FA")]),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
        ("ALIGN", (1,0), (-1,-1), "CENTER"),
    ]))
    elementos += [Paragraph("Resumo por tipo de serviço", estilos["Heading2"]), t, Spacer(1, 4*mm)]

    # Tabela de OS
    cab = ["OS", "Data", "Veículo", "Tipo de serviço", "Situação", "Início", "Fim", "Horas", "Custo"]
    linhas_tabela = [cab] + [[
        o["numero"], o["data"], o["veiculo"], o["grupo"] or o["tipo"],
        o["status"], o["hora_inicio"], o["hora_fim"], o["horas"], o["custo"]
    ] for o in dados["ordens"]]

    col_w = [18*mm, 22*mm, 30*mm, 38*mm, 24*mm, 16*mm, 16*mm, 16*mm, 22*mm]
    tabela = Table(linhas_tabela, colWidths=col_w, repeatRows=1)
    tabela.setStyle(TableStyle([
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1A6B8A")),
        ("TEXTCOLOR", (0,0), (-1,0), colors.white),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("FONTSIZE", (0,0), (-1,-1), 7.5),
        ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F4F7FA")]),
        ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
        ("ALIGN", (5,0), (-1,-1), "CENTER"),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("TOPPADDING", (0,0), (-1,-1), 3),
        ("BOTTOMPADDING", (0,0), (-1,-1), 3),
    ]))
    elementos += [Paragraph("Ordens de serviço executadas", estilos["Heading2"]), tabela]

    # Itens mais usados
    if dados["itens_frequentes"]:
        elementos.append(Spacer(1, 4*mm))
        cab_i = ["Peça / serviço", "Vezes usada", "Qtd total"]
        linhas_i = [cab_i] + [[i["descricao"], str(i["vezes"]), f"{i['qtd_total']:.0f}"] for i in dados["itens_frequentes"][:15]]
        t2 = Table(linhas_i, colWidths=[110*mm, 35*mm, 35*mm])
        t2.setStyle(TableStyle([
            ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#1A6B8A")),
            ("TEXTCOLOR", (0,0), (-1,0), colors.white),
            ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
            ("FONTSIZE", (0,0), (-1,-1), 8),
            ("ROWBACKGROUNDS", (0,1), (-1,-1), [colors.white, colors.HexColor("#F4F7FA")]),
            ("GRID", (0,0), (-1,-1), 0.3, colors.HexColor("#CCCCCC")),
            ("ALIGN", (1,0), (-1,-1), "CENTER"),
        ]))
        elementos += [Paragraph("Peças / serviços mais usados", estilos["Heading2"]), t2]

    doc.build(elementos)
    buf.seek(0)
    nome = f"mecanico_{mecanico.replace(' ','_')}_{hoje():%Y%m%d}.pdf"
    return send_file(buf, as_attachment=True, mimetype="application/pdf",
                     download_name=nome)


@bp_relatorios.get("/mecanico/<mecanico>.xlsx")
@visualizar_tela("relatorios")
def relatorio_mecanico_xlsx(mecanico):
    """Relatório de produtividade de um mecânico em Excel."""
    dados = _montar_dados_mecanico(mecanico)
    wb = Workbook()
    ws = wb.active
    ws.title = f"{mecanico[:20]}"

    cab_fill  = PatternFill("solid", fgColor="0F3D56")
    sub_fill  = PatternFill("solid", fgColor="1A6B8A")
    zebra     = PatternFill("solid", fgColor="F4F7FA")
    lado = openpyxl_side(style="thin", color="CCCCCC")
    borda = openpyxl_border(left=lado, right=lado, top=lado, bottom=lado)

    def _cel(ws, lin, col, val, bold=False, fundo=None, align="left", fmt=None):
        c = ws.cell(row=lin, column=col, value=val)
        c.font = openpyxl_font(bold=bold, size=10,
                               color="FFFFFF" if fundo in (cab_fill, sub_fill) else "000000")
        if fundo:
            c.fill = fundo
        c.alignment = openpyxl_align(horizontal=align, vertical="center")
        c.border = borda
        if fmt:
            c.number_format = fmt
        return c

    # Cabeçalho
    ws.merge_cells("A1:I1")
    c = ws.cell(row=1, column=1, value=f"Relatório de Serviços — {mecanico.title()}")
    c.font = openpyxl_font(bold=True, size=13, color="FFFFFF")
    c.fill = cab_fill
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:I2")
    c2 = ws.cell(row=2, column=1,
                 value=f"Período: {dados['inicio']:%d/%m/%Y} a {dados['fim']:%d/%m/%Y}  •  "
                       f"{dados['total_os']} OS  •  {dados['horas_totais']}h trabalhadas")
    c2.font = openpyxl_font(size=10, color="FFFFFF")
    c2.fill = sub_fill
    ws.row_dimensions[2].height = 18

    # Cabeçalho da tabela
    cabs = ["OS","Data","Veículo","Tipo de serviço","Situação","Início","Fim","Horas","Custo R$"]
    for i, cab in enumerate(cabs, 1):
        _cel(ws, 3, i, cab, bold=True, fundo=sub_fill, align="center")

    for idx, o in enumerate(dados["ordens"]):
        lin = idx + 4
        fundo = zebra if idx % 2 == 0 else None
        _cel(ws, lin, 1, o["numero"],           fundo=fundo, align="center")
        _cel(ws, lin, 2, o["data"],              fundo=fundo, align="center")
        _cel(ws, lin, 3, o["veiculo"],           fundo=fundo)
        _cel(ws, lin, 4, o["grupo"] or o["tipo"],fundo=fundo)
        _cel(ws, lin, 5, o["status"],            fundo=fundo, align="center")
        _cel(ws, lin, 6, o["hora_inicio"],       fundo=fundo, align="center")
        _cel(ws, lin, 7, o["hora_fim"],          fundo=fundo, align="center")
        _cel(ws, lin, 8, o["horas"],             fundo=fundo, align="center")
        _cel(ws, lin, 9, o["custo_num"],         fundo=fundo, align="right", fmt='R$ #,##0.00')

    for col, larg in zip("ABCDEFGHI", [12, 14, 26, 32, 18, 10, 10, 10, 14]):
        ws.column_dimensions[col].width = larg

    # Aba de peças mais usadas
    if dados["itens_frequentes"]:
        ws2 = wb.create_sheet("Peças mais usadas")
        for i, cab in enumerate(["Peça / serviço","Vezes usada","Qtd total"], 1):
            _cel(ws2, 1, i, cab, bold=True, fundo=sub_fill, align="center")
        for idx, it in enumerate(dados["itens_frequentes"], 2):
            fundo = zebra if idx % 2 == 0 else None
            _cel(ws2, idx, 1, it["descricao"], fundo=fundo)
            _cel(ws2, idx, 2, it["vezes"],     fundo=fundo, align="center")
            _cel(ws2, idx, 3, it["qtd_total"], fundo=fundo, align="center")
        ws2.column_dimensions["A"].width = 50
        ws2.column_dimensions["B"].width = 16
        ws2.column_dimensions["C"].width = 16

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    nome = f"mecanico_{mecanico.replace(' ','_')}_{hoje():%Y%m%d}.xlsx"
    return send_file(buf,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                     as_attachment=True, download_name=nome)


@bp_relatorios.get("/mecanicos.json")
@visualizar_tela("relatorios")
def listar_mecanicos_relatorio():
    """Lista mecânicos únicos para popular o select da tela de Relatórios."""
    from sqlalchemy import func
    nomes = (
        db.session.query(OrdemServico.mecanico)
        .filter(OrdemServico.mecanico.isnot(None), OrdemServico.mecanico != "")
        .distinct()
        .order_by(func.lower(OrdemServico.mecanico))
        .all()
    )
    vistos = {}
    for (nome,) in nomes:
        chave = nome.strip().upper()
        if chave not in vistos:
            vistos[chave] = nome.strip().title()
    return jsonify(sorted(vistos.values()))


def _montar_dados_mecanico(mecanico: str) -> dict:
    """Monta todos os dados de um mecânico para os relatórios."""
    from collections import Counter, defaultdict
    inicio, fim = _periodo()

    # OS do mecânico no período (busca por nome normalizado)
    todas_os = OrdemServico.query.filter(
        OrdemServico.data_abertura.between(inicio, fim)
    ).all()
    # Normaliza para achar "cleiton", "CLEITON", "Cleiton" como o mesmo
    ordens = [o for o in todas_os
              if (o.mecanico or "").strip().upper() == mecanico.strip().upper()]

    # Montar linhas
    linhas = []
    horas_total_min = 0
    for o in sorted(ordens, key=lambda x: x.data_abertura or hoje()):
        dur = o.duracao_minutos
        horas_str = f"{dur//60}h{dur%60:02d}m" if dur is not None else "—"
        if dur:
            horas_total_min += dur
        linhas.append({
            "numero":     o.numero or "—",
            "data":       o.data_abertura.strftime("%d/%m/%Y") if o.data_abertura else "—",
            "veiculo":    f"{o.veiculo.prefixo}/{o.veiculo.placa}" if o.veiculo else "—",
            "tipo":       o.tipo or "—",
            "grupo":      o.grupo or "",
            "status":     o.status or "—",
            "hora_inicio":o.hora_inicio.strftime("%H:%M") if o.hora_inicio else "—",
            "hora_fim":   o.hora_fim.strftime("%H:%M") if o.hora_fim else "—",
            "horas":      horas_str,
            "custo":      f"R$ {o.custo_total:,.2f}".replace(",","X").replace(".",",").replace("X","."),
            "custo_num":  o.custo_total,
        })

    # Resumo por tipo
    por_tipo_d = defaultdict(lambda: {"qtd":0,"min":0,"custo":0})
    for o in ordens:
        t = o.tipo or "Outros"
        por_tipo_d[t]["qtd"] += 1
        por_tipo_d[t]["min"] += o.duracao_minutos or 0
        por_tipo_d[t]["custo"] += o.custo_total
    por_tipo = [{
        "tipo": t, "qtd": v["qtd"],
        "horas": f"{v['min']//60}h{v['min']%60:02d}m",
        "custo": v["custo"]
    } for t, v in sorted(por_tipo_d.items())]

    # Itens mais usados
    contador = Counter()
    qtd_total = defaultdict(float)
    for o in ordens:
        for item in o.itens:
            desc = (item.descricao or "").strip()
            if desc:
                contador[desc] += 1
                qtd_total[desc] += item.quantidade or 0
    itens_freq = [{"descricao": k, "vezes": v, "qtd_total": qtd_total[k]}
                  for k, v in contador.most_common(20)]

    horas_tot = f"{horas_total_min//60}h{horas_total_min%60:02d}m"

    return {
        "mecanico":       mecanico.title(),
        "inicio":         inicio,
        "fim":            fim,
        "total_os":       len(ordens),
        "horas_totais":   horas_tot,
        "ordens":         linhas,
        "por_tipo":       por_tipo,
        "itens_frequentes": itens_freq,
    }


@bp_relatorios.get("/grupos-estoque.json")
@visualizar_tela("relatorios")
def grupos_estoque():
    """Grupos de peças cadastrados — alimenta o filtro de grupo dos relatórios."""
    grupos = [g for (g,) in db.session.query(Peca.grupo)
              .filter(Peca.grupo.isnot(None), Peca.grupo != "")
              .distinct().order_by(Peca.grupo).all()]
    return jsonify(grupos)


@bp_relatorios.post("/restaurar")
@perfil_obrigatorio("admin")
def restaurar_backup():
    """Recarrega os dados operacionais a partir de um backup JSON."""
    arquivo = request.files.get("arquivo")
    if not arquivo:
        return jsonify({"erro": "Selecione o arquivo de backup (.json)."}), 400
    if not arquivo.filename.lower().endswith(".json"):
        return jsonify({"erro": "O backup do SGMF é um arquivo .json."}), 400

    import json

    from services.crud import ErroNegocio
    try:
        pacote = json.load(arquivo.stream)
    except (json.JSONDecodeError, UnicodeDecodeError):
        return jsonify({"erro": "Não consegui ler o arquivo: ele não é um JSON válido."}), 400

    try:
        resumo = restaurar(pacote)
    except ErroNegocio as e:
        return jsonify({"erro": str(e)}), 400

    registrar_log("restaurar", "backup", 0, str(resumo))
    db.session.commit()
    return jsonify({"ok": True, "resumo": resumo,
                    "gerado_em": pacote.get("gerado_em")})


@bp_relatorios.get("/backup.json")
@visualizar_tela("relatorios")
def backup():
    """Cópia integral dos dados em JSON — útil antes de qualquer manutenção."""
    from models import Fornecedor, Motorista, Orcamento, ServicoTerceiro, Usuario
    pacote = {
        "gerado_em": agora().isoformat(),
        "veiculos": [v.to_dict() for v in Veiculo.query.all()],
        "motoristas": [m.to_dict() for m in Motorista.query.all()],
        "fornecedores": [f.to_dict() for f in Fornecedor.query.all()],
        "ordens": [o.to_dict(com_itens=True) for o in OrdemServico.query.all()],
        "servicos_terceiros": [s.to_dict() for s in ServicoTerceiro.query.all()],
        "abastecimentos": [a.to_dict() for a in Abastecimento.query.all()],
        "pneus": [p.to_dict() for p in Pneu.query.all()],
        "pecas": [p.to_dict() for p in Peca.query.all()],
        "movimentos": [m.to_dict() for m in MovimentoEstoque.query.all()],
        "orcamentos": [o.to_dict() for o in Orcamento.query.all()],
        "usuarios": [u.to_dict() for u in Usuario.query.all()],
    }
    resposta = jsonify(pacote)
    resposta.headers["Content-Disposition"] = \
        f"attachment; filename=sgmf_backup_{hoje():%Y%m%d}.json"
    return resposta
