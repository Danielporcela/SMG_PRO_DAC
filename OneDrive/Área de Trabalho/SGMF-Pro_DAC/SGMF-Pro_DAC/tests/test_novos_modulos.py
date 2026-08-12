"""Importação de planilhas, anexos, novos indicadores e aviso por e-mail."""
import io
from datetime import timedelta
from unittest.mock import patch

import pytest
from openpyxl import Workbook, load_workbook

from services.tempo import hoje as data_de_hoje


# ------------------------------------------------------------- importação
def planilha(cabecalhos, linhas):
    wb = Workbook()
    ws = wb.active
    ws.append(cabecalhos)
    for linha in linhas:
        ws.append(linha)
    arquivo = io.BytesIO()
    wb.save(arquivo)
    arquivo.seek(0)
    return arquivo


def enviar(cliente, tipo, arquivo, nome="planilha.xlsx"):
    return cliente.post(f"/api/importacao/{tipo}/conferir",
                        data={"arquivo": (arquivo, nome)},
                        content_type="multipart/form-data")


@pytest.mark.parametrize("tipo", ["veiculos", "motoristas", "fornecedores", "pecas"])
def test_modelo_da_planilha_pode_ser_baixado(logado, tipo):
    r = logado.get(f"/importacao/modelo/{tipo}.xlsx")
    assert r.status_code == 200
    assert "spreadsheetml" in r.headers["Content-Type"]
    wb = load_workbook(io.BytesIO(r.data))
    assert wb.active.max_row >= 2, "o modelo traz cabeçalho e uma linha de exemplo"


def test_modelo_marca_as_colunas_obrigatorias(logado):
    wb = load_workbook(io.BytesIO(logado.get("/importacao/modelo/veiculos.xlsx").data))
    cabecalhos = [c.value for c in wb.active[1]]
    assert "Prefixo *" in cabecalhos and "Placa *" in cabecalhos
    assert "Marca" in cabecalhos


def test_importacao_de_veiculos(logado):
    arquivo = planilha(["Prefixo", "Placa", "Marca", "Hodômetro"],
                       [["FR-201", "imp1a11", "Volvo", 50000],
                        ["FR-202", "imp2b22", "Scania", 61000]])
    previa = enviar(logado, "veiculos", arquivo).get_json()
    assert previa["total"] == 2
    assert len(previa["prontas"]) == 2
    assert previa["problemas"] == []
    assert previa["prontas"][0]["dados"]["placa"] == "IMP1A11", "placa vem em maiúsculas"

    r = logado.post("/api/importacao/veiculos/gravar", json={"linhas": previa["prontas"]})
    assert r.status_code == 200 and r.get_json()["gravadas"] == 2
    prefixos = [v["prefixo"] for v in logado.get("/api/veiculos").get_json()]
    assert "FR-201" in prefixos and "FR-202" in prefixos


def test_conferencia_nao_grava_nada(logado):
    arquivo = planilha(["Nome"], [["Motorista da Prévia"]])
    enviar(logado, "motoristas", arquivo)
    assert logado.get("/api/motoristas").get_json() == []


def test_linha_sem_campo_obrigatorio_e_recusada(logado):
    arquivo = planilha(["Prefixo", "Placa"], [["FR-300", ""], ["FR-301", "OKA1B22"]])
    previa = enviar(logado, "veiculos", arquivo).get_json()
    assert len(previa["prontas"]) == 1
    assert len(previa["problemas"]) == 1
    assert "obrigatório" in previa["problemas"][0]["erro"]
    assert previa["problemas"][0]["linha"] == 2


def test_numero_invalido_e_apontado_com_a_linha(logado):
    arquivo = planilha(["Prefixo", "Placa", "Hodômetro"],
                       [["FR-400", "NUM1A11", "muito rodado"]])
    previa = enviar(logado, "veiculos", arquivo).get_json()
    assert previa["problemas"][0]["linha"] == 2
    assert "Hodômetro" in previa["problemas"][0]["erro"]


def test_placa_ja_cadastrada_e_recusada(logado, base):
    arquivo = planilha(["Prefixo", "Placa"], [["FR-500", "ABC1D23"]])
    previa = enviar(logado, "veiculos", arquivo).get_json()
    assert "já existe" in previa["problemas"][0]["erro"]


def test_placa_repetida_na_propria_planilha(logado):
    arquivo = planilha(["Prefixo", "Placa"],
                       [["FR-600", "REP1A11"], ["FR-601", "REP1A11"]])
    previa = enviar(logado, "veiculos", arquivo).get_json()
    assert len(previa["prontas"]) == 1
    assert "repetido na própria planilha" in previa["problemas"][0]["erro"]


def test_linhas_em_branco_sao_ignoradas(logado):
    arquivo = planilha(["Nome"], [["Motorista A"], [None], [""], ["Motorista B"]])
    previa = enviar(logado, "motoristas", arquivo).get_json()
    assert previa["total"] == 2


def test_planilha_sem_as_colunas_certas(logado):
    arquivo = planilha(["Carro", "Chapa"], [["x", "y"]])
    r = enviar(logado, "veiculos", arquivo)
    assert r.status_code == 400
    assert "Prefixo" in r.get_json()["erro"]


def test_arquivo_que_nao_e_planilha(logado):
    r = enviar(logado, "veiculos", io.BytesIO(b"isto nao e planilha"), "coisa.txt")
    assert r.status_code == 400


def test_importacao_de_pecas_cria_o_saldo_inicial(logado):
    arquivo = planilha(["Código", "Descrição", "Saldo inicial", "Custo unitário"],
                       [["IMP-01", "Filtro importado", 15, 30]])
    previa = enviar(logado, "pecas", arquivo).get_json()
    logado.post("/api/importacao/pecas/gravar", json={"linhas": previa["prontas"]})

    peca = logado.get("/api/pecas").get_json()[0]
    assert peca["quantidade"] == 15
    assert peca["custo_unitario"] == 30
    movimentos = logado.get("/api/movimentos").get_json()
    assert movimentos[0]["tipo"] == "entrada"
    assert "Importação" in movimentos[0]["documento"]


def test_datas_da_planilha_sao_lidas(logado):
    arquivo = planilha(["Nome", "Validade da CNH"], [["Motorista Data", "2028-03-15"]])
    previa = enviar(logado, "motoristas", arquivo).get_json()
    logado.post("/api/importacao/motoristas/gravar", json={"linhas": previa["prontas"]})
    assert logado.get("/api/motoristas").get_json()[0]["validade_cnh"] == "2028-03-15"


def test_importacao_registra_na_auditoria(logado):
    arquivo = planilha(["Nome"], [["Motorista Auditado"]])
    previa = enviar(logado, "motoristas", arquivo).get_json()
    logado.post("/api/importacao/motoristas/gravar", json={"linhas": previa["prontas"]})
    logs = logado.get("/api/logs").get_json()
    assert any(l["acao"] == "importar" for l in logs)


def test_consulta_nao_importa(logado, cliente):
    logado.post("/api/usuarios", json={"nome": "Fiscal", "email": "fi@teste.local",
                                       "perfil": "consulta", "senha": "fiscal123"})
    logado.get("/logout")
    cliente.post("/login", json={"email": "fi@teste.local", "senha": "fiscal123"})
    assert cliente.get("/importacao/modelo/veiculos.xlsx").status_code == 403
    arquivo = planilha(["Nome"], [["Alguém"]])
    assert enviar(cliente, "motoristas", arquivo).status_code == 403


# ----------------------------------------------------------------- anexos
def arquivo_falso(nome="nota.pdf", tipo="application/pdf", tamanho=2048):
    return (io.BytesIO(b"%PDF-1.4" + b"x" * tamanho), nome)


def anexar(cliente, tipo, registro_id, arquivo=None, descricao="Nota fiscal"):
    return cliente.post(f"/api/anexos/{tipo}/{registro_id}",
                        data={"arquivo": arquivo or arquivo_falso(), "descricao": descricao},
                        content_type="multipart/form-data")


@pytest.fixture()
def ordem(logado, base):
    return logado.post("/api/ordens", json={"veiculo_id": base["veiculo"]["id"]}).get_json()


def test_anexar_arquivo_na_os(logado, ordem):
    r = anexar(logado, "ordens", ordem["id"])
    assert r.status_code == 201
    anexo = r.get_json()
    assert anexo["nome"] == "nota.pdf"
    assert anexo["descricao"] == "Nota fiscal"
    assert anexo["enviado_por"] == "Administrador"
    assert anexo["imagem"] is False


def test_lista_e_download_do_anexo(logado, ordem):
    criado = anexar(logado, "ordens", ordem["id"]).get_json()
    lista = logado.get(f"/api/anexos/ordens/{ordem['id']}").get_json()
    assert len(lista) == 1

    arquivo = logado.get(f"/api/anexos/{criado['id']}/arquivo")
    assert arquivo.status_code == 200
    assert arquivo.data.startswith(b"%PDF")
    assert "application/pdf" in arquivo.headers["Content-Type"]


def test_contador_de_anexos_aparece_na_os(logado, ordem):
    anexar(logado, "ordens", ordem["id"])
    assert logado.get(f"/api/ordens/{ordem['id']}").get_json()["qtd_anexos"] == 1


def test_anexo_em_abastecimento(logado, base):
    abastecimento = logado.post("/api/abastecimentos", json={
        "veiculo_id": base["veiculo"]["id"], "km_atual": 100200,
        "litros": 50, "valor_litro": 6}).get_json()
    r = anexar(logado, "abastecimentos", abastecimento["id"],
               arquivo=(io.BytesIO(b"\x89PNG" + b"y" * 500), "cupom.png"),
               descricao="Cupom do posto")
    assert r.status_code == 201
    assert r.get_json()["imagem"] is True


def test_arquivo_grande_e_recusado(logado, ordem, app):
    grande = (io.BytesIO(b"%PDF" + b"z" * (6 * 1024 * 1024)), "grande.pdf")
    r = anexar(logado, "ordens", ordem["id"], arquivo=grande)
    assert r.status_code == 400
    assert "limite" in r.get_json()["erro"]


def test_tipo_de_arquivo_nao_permitido(logado, ordem):
    r = anexar(logado, "ordens", ordem["id"],
               arquivo=(io.BytesIO(b"MZ executavel"), "virus.exe"))
    assert r.status_code == 400
    assert "foto" in r.get_json()["erro"].lower()


def test_anexo_sem_arquivo(logado, ordem):
    r = logado.post(f"/api/anexos/ordens/{ordem['id']}", data={},
                    content_type="multipart/form-data")
    assert r.status_code == 400


def test_excluir_anexo(logado, ordem):
    criado = anexar(logado, "ordens", ordem["id"]).get_json()
    assert logado.delete(f"/api/anexos/{criado['id']}").status_code == 200
    assert logado.get(f"/api/anexos/ordens/{ordem['id']}").get_json() == []


def test_excluir_a_os_leva_os_anexos_junto(logado, ordem):
    criado = anexar(logado, "ordens", ordem["id"]).get_json()
    logado.delete(f"/api/ordens/{ordem['id']}")
    assert logado.get(f"/api/anexos/{criado['id']}/arquivo").status_code == 404


def test_anexo_em_registro_inexistente(logado):
    assert anexar(logado, "ordens", 9999).status_code == 404


def test_consulta_ve_mas_nao_anexa(logado, cliente, ordem):
    anexar(logado, "ordens", ordem["id"])
    logado.post("/api/usuarios", json={"nome": "Fiscal", "email": "fi2@teste.local",
                                       "perfil": "consulta", "senha": "fiscal123"})
    logado.get("/logout")
    cliente.post("/login", json={"email": "fi2@teste.local", "senha": "fiscal123"})
    assert len(cliente.get(f"/api/anexos/ordens/{ordem['id']}").get_json()) == 1
    assert anexar(cliente, "ordens", ordem["id"]).status_code == 403


# --------------------------------------------------- indicadores novos
def test_prazo_medio_de_atendimento(logado, base):
    vid = base["veiculo"]["id"]
    for dias in (2, 6):
        abertura = (data_de_hoje() - timedelta(days=dias)).isoformat()
        os_criada = logado.post("/api/ordens", json={"veiculo_id": vid,
                                                     "data_abertura": abertura}).get_json()
        logado.put(f"/api/ordens/{os_criada['id']}", json={"status": "Finalizada"})

    d = logado.get("/api/painel/resumo").get_json()
    assert d["prazo_medio_atendimento"] == 4.0     # (2 + 6) / 2
    assert d["os_finalizadas"] == 2


def test_prazo_medio_ignora_os_em_aberto(logado, base):
    logado.post("/api/ordens", json={"veiculo_id": base["veiculo"]["id"]})
    d = logado.get("/api/painel/resumo").get_json()
    assert d["prazo_medio_atendimento"] is None
    assert d["os_abertas"] == 1


def test_economia_sem_historico_fica_indefinida(logado, base):
    d = logado.get("/api/painel/resumo").get_json()
    assert d["economia_periodo"] is None
    assert d["custo_km_historico"] is None


def test_economia_quando_o_custo_por_km_melhora(logado, base):
    """Histórico caro, período atual barato: precisa apontar economia."""
    vid = base["veiculo"]["id"]
    antigo = data_de_hoje() - timedelta(days=120)
    # histórico: 1.000 km gastando R$ 2.000 -> R$ 2,00/km
    logado.post("/api/abastecimentos", json={"veiculo_id": vid, "km_atual": 100000,
                                             "litros": 100, "valor_total": 1000,
                                             "data": antigo.isoformat()})
    logado.post("/api/abastecimentos", json={"veiculo_id": vid, "km_atual": 101000,
                                             "litros": 100, "valor_total": 2000,
                                             "data": (antigo + timedelta(days=2)).isoformat()})
    # agora: 1.000 km gastando R$ 1.000 -> R$ 1,00/km
    logado.post("/api/abastecimentos", json={"veiculo_id": vid, "km_atual": 102000,
                                             "litros": 100, "valor_total": 1000})

    # histórico: R$ 3.000 gastos para 1.000 km medidos -> R$ 3,00/km
    # agora: R$ 1.000 para 1.000 km -> R$ 1,00/km
    d = logado.get("/api/painel/resumo").get_json()
    assert d["custo_km_historico"] == 3.0
    assert d["custo_por_km"] == 1.0
    assert d["economia_periodo"] == 2000.0        # (3,00 - 1,00) x 1.000 km
    assert d["variacao_custo_km"] == -66.7


# ------------------------------------------------------ aviso por e-mail
def test_situacao_mostra_que_o_envio_esta_desligado(logado):
    d = logado.get("/api/notificacoes/situacao").get_json()
    assert d["configurado"] is False
    assert d["remetente"] == "estudosti20@gmail.com"
    assert "estudosti20@gmail.com" in d["destinatarios"]
    assert d["servidor"].startswith("smtp.gmail.com")


def test_teste_de_envio_sem_senha_explica_o_motivo(logado):
    r = logado.post("/api/notificacoes/testar", json={})
    assert r.status_code == 400
    assert "SMTP_SENHA" in r.get_json()["erro"]


def test_envio_de_teste_com_servidor_simulado(app, logado):
    app.config.update(SMTP_SENHA="senha-de-app")
    with patch("services.notificacoes.smtplib.SMTP") as servidor:
        r = logado.post("/api/notificacoes/testar",
                        json={"destinatario": "chefe@empresa.com.br"})
    assert r.status_code == 200
    assert r.get_json()["destinatarios"] == ["chefe@empresa.com.br"]
    servidor.assert_called_once()
    conexao = servidor.return_value          # o código usa "with servidor:" sem "as"
    conexao.login.assert_called_once_with("estudosti20@gmail.com", "senha-de-app")
    conexao.send_message.assert_called_once()


def test_senha_recusada_vira_mensagem_clara(app, logado):
    import smtplib
    app.config.update(SMTP_SENHA="errada")
    with patch("services.notificacoes.smtplib.SMTP") as servidor:
        servidor.return_value.login.side_effect = \
            smtplib.SMTPAuthenticationError(535, b"bad")
        r = logado.post("/api/notificacoes/testar", json={})
    assert r.status_code == 400
    assert "senha de aplicativo" in r.get_json()["erro"]


def test_resumo_diario_com_alertas(app, logado, base):
    app.config.update(SMTP_SENHA="senha-de-app")
    logado.put(f"/api/veiculos/{base['veiculo']['id']}", json={"hodometro": 130000})

    with patch("services.notificacoes.smtplib.SMTP") as servidor:
        r = logado.post("/api/notificacoes/enviar-agora", json={})
    dados = r.get_json()
    assert dados["enviado"] is True
    assert dados["alertas"] >= 1
    mensagem = servidor.return_value.send_message.call_args[0][0]
    html = mensagem.get_body(preferencelist=("html",)).get_content()
    assert "troca de óleo vencida" in html
    assert "Ação imediata" in html
    assert mensagem["To"] == "estudosti20@gmail.com"


def test_resumo_nao_envia_quando_esta_tudo_em_ordem(app, logado):
    app.config.update(SMTP_SENHA="senha-de-app")
    with patch("services.notificacoes.smtplib.SMTP") as servidor:
        r = logado.post("/api/notificacoes/enviar-agora", json={})
    assert r.get_json()["enviado"] is False
    assert "Nenhum alerta" in r.get_json()["motivo"]
    servidor.assert_not_called()


def test_aviso_so_sai_uma_vez_por_dia(app, logado, base):
    from services.notificacoes import executar_alertas_diarios

    app.config.update(SMTP_SENHA="senha-de-app")
    logado.put(f"/api/veiculos/{base['veiculo']['id']}", json={"hodometro": 130000})

    with app.app_context(), patch("services.notificacoes.smtplib.SMTP"):
        primeiro = executar_alertas_diarios()
        segundo = executar_alertas_diarios()
    assert primeiro["enviado"] is True
    assert segundo["enviado"] is False
    assert "já foi enviado" in segundo["motivo"]


def test_url_da_tarefa_exige_chave(app, cliente):
    assert cliente.get("/tarefas/alertas-diarios").status_code == 403
    app.config.update(CHAVE_TAREFAS="chave-secreta")
    assert cliente.get("/tarefas/alertas-diarios?chave=errada").status_code == 403
    assert cliente.get("/tarefas/alertas-diarios?chave=chave-secreta").status_code == 200


def test_notificacoes_sao_so_do_administrador(logado, cliente):
    logado.post("/api/usuarios", json={"nome": "Operador", "email": "op9@teste.local",
                                       "perfil": "operador", "senha": "operador123"})
    logado.get("/logout")
    cliente.post("/login", json={"email": "op9@teste.local", "senha": "operador123"})
    assert cliente.get("/api/notificacoes/situacao").status_code == 403
    assert cliente.post("/api/notificacoes/testar", json={}).status_code == 403


@pytest.mark.parametrize("rota", ["/importacao", "/notificacoes"])
def test_paginas_novas_abrem(logado, rota):
    assert logado.get(rota).status_code == 200
